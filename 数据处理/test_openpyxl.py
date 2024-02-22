import openpyxl
import xlrd
import xlwt
import random
#参考url：https://www.cnblogs.com/poloyy/p/12241448.html
data_path = r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\test_openpyxl.xlsx"


def writeExcel():
    # 创建工作簿对象，就是创建一个空的excel表；
    wb = openpyxl.Workbook()
    # 在索引为0的位置创建一个名为mytest的sheet页；创建多个sheet页面；
    sheet = wb.create_sheet("mytest", 0)
    sheet2 = wb.create_sheet("mytest1", 1)
    sheet3 = wb.create_sheet("mytest2", 2)
    # 将创建的工作簿保存为mytest.xlsx
    wb.save(r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\auto_create.xlsx")
    '''
    1、写入数据的文件需要是可写文件，需要已关闭；若开着excel则无法写入会报错；PermissionError: [Errno 13] Permission denied
    2、主要是是四部曲，获得Excel文件；执行workbook.active；赋值操作；保存文件；
    3、赋值操作还可以是 sheet['B2']=‘123’,等价于sheet.cell()
    4、save()，会覆盖原有文件，不会有提醒；
    '''
    workbook=openpyxl.load_workbook(r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\auto_create.xlsx")
    #当前打开的sheet页，wb.active;
    sheet=workbook.active
    #修改sheet页的名称；
    sheet.title="Update_mytest"
    #sheet页WS的tab颜色变更
    sheet.sheet_properties.tabColor="1072BA"
    #通过循环批量写入；
    for i in range(1,20):
        for j in range(1,15):
            sheet.cell(row=i,column=j,value=random.randint(1,100))
    # sheet.cell(row=1,column=1,value="write_value")
    #直接给某个单元格复制
    sheet['A1']='单独赋值'
    #指定行列的赋值
    sheet.cell(row=6,column=6,value="指定行列赋值")
    '''
    一次输出行，通过列表的形式一次性填入到excel表；
    '''
    #列名
    column_title=['Firstname','LastName']
    rows=[
        column_title,['1111','1111'],['222','222'],['333','333'],['444','444'],
    ]
    for row in rows:
        sheet.append(row)
    '''
    #插入行，在第7行之前插入
    ws.insert_rows(7)
    #插入列，在第7列之前插入
    ws.insert_cols(7)
    #删除行列
    ws.delete_rows(7)
    ws.delete_cols(7)
    #可以删除多个
    ws.delete_cols(6, 3)
    合并与拆分单元格
    #合并单元格,以最左上角写入数据或读取数据
    ws.merge_cells('A2:D2')
    #拆分单元格
    ws.unmerge_cells('A2:D2')
    '''
    #插入行，在第7行之前插入
    sheet.insert_rows(7)

    #保存excel表；
    workbook.save(r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\auto_create.xlsx")
    # 最后关闭文件
    wb.close()

writeExcel()

def readExcel():
    # 打开文件,load_workbook()等同于调用open();
    wb = openpyxl.load_workbook(data_path)
    # 获得所有sheet页的值，得到的是一个列表；
    sheets = wb.sheetnames
    for sheet in sheets:
        print(sheet)
    worksheet1 = wb[sheets[0]]
    worksheet2 = wb.get_sheet_by_name(sheets[1])
    print(worksheet1, worksheet2)
    # cell(row,colum,value=None)，三个值分别是行，列，值，若设置了value相当于赋值操作，会覆盖原本的值；
    cell_val = worksheet1.cell(row=2, column=2).value
    print('得到二行二列的值：', cell_val)
    print('获得A4的值：', worksheet1['A4'].value)
    print("获得A1-B3单元格的值共六个：", worksheet1['A1':'B3'])
    a1b3 = worksheet1['A1':'B3']
    for v in a1b3:
        print('获得对应的值：', v[1].value)
    # 获得第二行的值
    a2 = worksheet1[2]
    print('获得第二行的值', len(a2), a2)
    for v in range(len(a2)):
        print('第二行的第' + str(v) + '个值', a2[v].value)
    # 获得第1-3行单元格
    print('获得第1-3行单元格', len(worksheet1[1:3]), worksheet1[1:3])
    print("******************************************")
    '''
    sheets.rows，返回的是一个对象，需要用tuple()，才能将对象转换成tuple；
    '''
    print('以列的形式，获取sheet的全部cell：',worksheet1.columns,'\n',tuple(worksheet1.columns))
    print('以行的形式，获取sheet的全部cell：',worksheet1.rows,'\n',tuple(worksheet1.rows))
    #获取所有数据
    all_data=tuple(worksheet1.values)
    print('获取sheet页的所有数据：',all_data)
    #指定返回某一行数据，行数是从1开始的；
    print('获取第二行的所有数据',all_data[2],'\n',list(all_data[2]))
    print('获取第二行第二个数据', all_data[2][1])
    #获取sheet的行数和列数
    print("获取sheet的最大行：",worksheet1.max_row)
    print("获取sheet的最小行：", worksheet1.min_row)
    print("获取sheet的最大列：", worksheet1.max_column)
    print("获取sheet的最小列：", worksheet1.min_column)
    '''
    通过遍历获得文件的所有数据，并把它们放入列表中；
    '''
    mrow=worksheet1.max_row
    mcolumn=worksheet1.max_column
    for i in range(2,mrow+1):
        l="list"+str(i)
        l=[]
        for j in range(1,mcolumn+1):
            l.append(worksheet1.cell(row=i,column=j).value)
        print('第'+str(i)+'的值是：',l)
            # print('获得excel的值是：',worksheet1.cell(row=i,column=j).value)

    '''
    获取某列的值，
    '''
    c_str="ABCDEFGHIJKLMN"
    for i in range(len(c_str)):
        # print(c_str[i])
        l = "list" + str(c_str[i])
        l = []
        for j in worksheet1[c_str[i]]:
            l.append(j.value)
        print('第 '+str(c_str[i])+' 列的值是：',l)
    #行列生成器，sheet.rows/columns，分别为行/列生成器，都是每一行/列的cell对象，有一个tuple包裹；
    print(worksheet1.rows,worksheet2.rows)
    for row in worksheet1.rows:
        for cell in row:
            print(cell.value)









# readExcel()









