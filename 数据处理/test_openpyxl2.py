import openpyxl
import xlrd
import xlwt
import random


def writeExcel():
    # 创建工作簿对象，就是创建一个空的excel表；
    wb = openpyxl.Workbook()
    # 在索引为0的位置创建一个名为mytest的sheet页；创建多个sheet页面；
    sheet = wb.create_sheet("mytest", 0)
    sheet2 = wb.create_sheet("mytest1", 1)
    print('当前sheet的所有sheet页：',wb.sheetnames)
    ws=wb.active
    ws['A1']='单独写入'
    ws['A2'] = '单独写入'
    #复制一个sheet对象
    target=wb.copy_worksheet(ws)
    print(target.values)
    #删除sheet
    wb.remove(target)



    # 将创建的工作簿保存为mytest.xlsx
    wb.save(r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\auto_create2.xlsx")

writeExcel()