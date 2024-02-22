import xlrd
import xlwt
import random
import openpyxl

# 参考url：https://www.cnblogs.com/ubuntu1987/p/11933376.html
data_path = r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\test_xlrd.xls"

'''对xls，xlsx文件进行读操作；'''


def readXlrd():
    # 打开excel文件读取数据；
    wb = xlrd.open_workbook(data_path)
    # 获取工作表sheet对象
    sheets = wb.sheet_names()
    print("以列表的形式获得所有sheet页名字：", sheets)
    # 可以根据sheet页名字获取指定表名的表，返回一个对象；也可以根据索引获取对应sheet表，索引从0开始，返回一个对象；
    sheet = wb.sheet_by_name(sheets[0])
    sheet = wb.sheet_by_index(0)
    sheetName = sheet.name
    print('获取工作表对象以及sheet名', sheetName, sheet)
    # 获得sheet的行数和列数
    print('sheet表的总行数：', sheet.nrows)
    print('sheet表总列数：', sheet.ncols)
    # 获取整行或者整列的值（数值），row_values,col_values，索引从0开始；
    row_values = sheet.row_values(1)
    print('获取第一行的值：', row_values)
    column_values = sheet.col_values(1)
    print('获取第一列的值：', column_values)
    # 获取指定单元格的值，cell(a,b).value,row(1)[0].value，返回的是一个列表；
    print('获取指定单元格的值：', sheet.cell(1, 1).value)
    print('获取二行二列的值：', sheet.row(1)[1].value)
    # 获取单元格内容的数据类型，返回代表数据类型的值，编码分别代表：ctype : 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
    print('单元格值的数据类型：', sheet.cell(1, 1).ctype)


# readXlrd()
'''对xls文件进行写操作；'''


def writeXlwt():
    wb = xlwt.Workbook(encoding='utf-8')  # 创建一个工作簿；
    sheet = wb.add_sheet("xlwt写入数据")
    # 向工作簿中下入数据
    for i in range(20):
        for j in range(20):
            sheet.write(i, j, random.randint(1, 100))
    # 保存工作簿
    wb.save(data_path)


writeXlwt()

'''对xlsx文件进行写操作'''
d_path = r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\test_openpyxl.xlsx"


def writeOpenpyxl():
    # 新建工作簿
    wb = openpyxl.Workbook()
    # 在工作簿中新建sheet页
    sheet = wb.create_sheet("operateOpenpyxl", 0)
    # 向表格中写入数据，cell(i,j,value)
    # 获得当前活跃的工作页，默认是第一个工作页
    sheetactive = wb.active
    # 向i行j列写入数据value，注意，行号和列号都是从1开始计数；
    for i in range(1, 10):
        for j in range(1, 12):
            sheetactive.cell(i, j, value=random.randint(1, 100))
    wb.save(d_path)


writeOpenpyxl()
'''对xlsx进行读操作'''


def readOpenpyxl():
    # 读取数据，注：filemname为文件名以及路径，如果路径或者文件名有中文给前面加一个r表示原生字符。
    wb = openpyxl.load_workbook(d_path)
    print('获取所有sheet页名字', wb.sheetnames)
    # 根据sheet页的名字获取指定表名的表；
    sheetName = wb.get_sheet_by_name('Sheet')
    # 根据索引获取对一个sheet的表；
    sheetName = wb.worksheets[0]
    print('获得sheet的名称：', sheetName.title)
    print('获得总行/列数：', sheetName.max_row, sheetName.max_column)
    # 获取整行或者整列的值,获得是一个生成器，里面是每一行的数据，每一行数据由一个元祖类型包裹；
    row_values = sheetName.rows
    print(row_values)
    # for row in row_values:
    #     for cell in row:
    #         print(cell.value)
    print('获取每一列的值：', sheetName.columns, tuple(sheetName.columns))
    for cell in sheetName.columns:
        for ce in cell:
            print(ce.value)
    print("获取第一行的值，", sheetName.rows)
    # 获取单元格的值，cell(a,b).value,索引从1开始
    # 注：此处的行数和列数都是从1开始计数的，而在xlrd中是由0开始计数的
    print(sheetName['A1'].value)
    print(sheetName.cell(1, 1).value)


readOpenpyxl()

'''openpyxl向指定单元格中添加图片并修改图片大小'''
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os

img_path = r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\test.png"
img_path2 = r"C:\Users\zhangbinghui\PycharmProjects\anaconda\数据处理\test2.png"

def insertOperpyxl():
    wb = Workbook()
    sheet = wb.active
    # 设置文字图片单元格的高列宽，设置行高，该设置行高和excel文件中的行高是一样的；
    # column_width = 10
    # row_height = 80
    img = Image(img_path)
    img2 = Image(img_path2)
    newSize = (90, 90)
    img.width, img.height = newSize
    img2.width, img2.height = newSize
    # 向d列中的单元格内指定添加图片
    sheet.add_image(img, 'd' + str(6))
    sheet.add_image(img2, 'a' + str(6))
    wb.save(d_path)
insertOperpyxl()
